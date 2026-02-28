import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path

# ==========================================================
# BASE PATH (ONLY CHANGE)
# ==========================================================
BASE_DIR = Path(__file__).resolve().parent

PATH_28  = BASE_DIR / "cleaning_file.xlsx"
PATH_REV = BASE_DIR / "revision_extraction.xlsx"
OUT_PATH = BASE_DIR / "raw_validated.xlsx"

# ==========================================================
# LOAD EXCELS
# ==========================================================

df_28  = pd.read_excel(PATH_28, dtype=str, keep_default_na=False)
df_rev = pd.read_excel(PATH_REV, dtype=str, keep_default_na=False)

# ==========================================================
# BLAD FIXING
# ==========================================================

df_28["BLAD"] = (
    df_28["BLAD"]
    .astype(str)
    .str.strip()
    .str.zfill(3)
)

# ==========================================================
# CREATING COMMON KEY
# ==========================================================

df_28["DOC_KEY"] = (
    df_28["Image"]
    .str.replace("_stamp.png", "", regex=False)
    .str.strip()
)

df_rev["DOC_KEY"] = (
    df_rev["FILE"]
    .str.replace(r"_p\d+", "", regex=True)
    .str.replace(r"\.(png|pdf)$", "", regex=True, case=False)
    .str.strip()
)

# ==========================================================
# MERGE
# ==========================================================

df_final = df_28.merge(
    df_rev[["DOC_KEY", "FINAL_REV"]],
    on="DOC_KEY",
    how="left"
)

df_final.drop(columns=["DOC_KEY"], inplace=True)

df_final["REV_STATUS"] = "OK"

# ==========================================================
# HELPERS
# ==========================================================

def normalize_revision(val: str) -> str:
    if not val:
        return ""
    return val.strip().upper()

def is_valid_revision(val: str) -> bool:
    return bool(re.fullmatch(r"[A-Z](?:\.\d+)?", val))

def is_pure_number(val: str) -> bool:
    return bool(val and val.strip().isdigit())

# ==========================================================
# REVISION LOGIC
# ==========================================================

for idx, row in df_final.iterrows():

    andr = normalize_revision(str(row["ANDR"])) if row["ANDR"] else ""
    rev  = normalize_revision(str(row["FINAL_REV"])) if row["FINAL_REV"] else ""

    andr_valid = is_valid_revision(andr)
    rev_valid  = is_valid_revision(rev)

    if is_pure_number(andr):
        continue

    if andr_valid and rev_valid and andr != rev:
        df_final.at[idx, "REV_STATUS"] = "ERROR"

    elif andr_valid and not rev_valid:
        df_final.at[idx, "REV_STATUS"] = "ERROR"

    elif rev_valid and not andr_valid:
        df_final.at[idx, "REV_STATUS"] = "ERROR"

# ==========================================================
# SAVE TO EXCEL
# ==========================================================

df_final.to_excel(OUT_PATH, index=False)

wb = load_workbook(OUT_PATH)
ws = wb.active

red_fill = PatternFill(
    start_color="FFFF0000",
    end_color="FFFF0000",
    fill_type="solid"
)

headers = {
    str(cell.value).strip(): idx + 1
    for idx, cell in enumerate(ws[1])
}

andr_col   = headers["ANDR"]
rev_col    = headers["FINAL_REV"]
status_col = headers["REV_STATUS"]

for row in range(2, ws.max_row + 1):

    status = ws.cell(row=row, column=status_col).value

    if status == "ERROR":
        ws.cell(row=row, column=andr_col).fill = red_fill
        ws.cell(row=row, column=rev_col).fill  = red_fill

wb.save(OUT_PATH)

print("=" * 60)
print("DONE ")
print("REV_STATUS column added")
print("• Errors stored in DATA (not colors)")
print("• Stable multi-step pipeline")
print("• ANDR vs FINAL_REV mismatches marked")
print(OUT_PATH)
print("=" * 60)