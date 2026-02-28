import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path

# ============================================================
# BASE PATH (ONLY CHANGE)
# ============================================================
BASE_DIR = Path(__file__).resolve().parent

DATA_FILE   = BASE_DIR / "raw_validated.xlsx"
MASTER_FILE = BASE_DIR / "mastercopy_labels.xlsx"
OUTPUT_FILE = BASE_DIR / "validation_file.xlsx"

# ============================================================
# LOAD DATA
# ============================================================
df = pd.read_excel(DATA_FILE, dtype=str, keep_default_na=False)

master_value    = pd.read_excel(MASTER_FILE, sheet_name="VALUE",    dtype=str, keep_default_na=False)
master_pattern  = pd.read_excel(MASTER_FILE, sheet_name="PATTERN",  dtype=str, keep_default_na=False)
master_freetext = pd.read_excel(MASTER_FILE, sheet_name="FREETEXT", dtype=str, keep_default_na=False)

# ============================================================
# CLEAN MASTER TABLES
# ============================================================
def clean_master(df):
    df["LABEL"] = df["LABEL"].astype(str).str.strip()
    if "VALUE" in df.columns:
        df["VALUE"] = df["VALUE"].astype(str).str.strip()
    return df

master_value    = clean_master(master_value)
master_pattern  = clean_master(master_pattern)
master_freetext = clean_master(master_freetext)

# ============================================================
# RULE LOOKUPS
# ============================================================
value_rules   = master_value.groupby("LABEL")["VALUE"].apply(set).to_dict()
pattern_rules = master_pattern.groupby("LABEL")["VALUE"].apply(list).to_dict()
freetext_labels = set(master_freetext["LABEL"])

EMPTY_ALLOWED = {
    "BANDEL","BLAD","NASTA_BLAD","KILOMETER_METER","ANDR",
    "ANLAGGNINGSTYP","GRANSKNINGSSTATUS_SYFTE","HANDLINGSTYP",
    "SKALA","FORMAT","DATUM","TEKNIKOMRADE"
}

# ============================================================
# WRITE DATA FIRST
# ============================================================
df.to_excel(OUTPUT_FILE, index=False)

wb = load_workbook(OUTPUT_FILE)
ws = wb.active

RED = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

headers = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1])}

andr_col   = headers["ANDR"]
rev_col    = headers["FINAL_REV"]
status_col = headers["REV_STATUS"]

# ============================================================
# REVISION ERRORS
# ============================================================
for row_idx in range(2, len(df) + 2):

    status = ws.cell(row=row_idx, column=status_col).value

    if status == "ERROR":
        ws.cell(row=row_idx, column=andr_col).fill = RED
        ws.cell(row=row_idx, column=rev_col).fill  = RED

# ============================================================
# RNP SUBSET OF IMAGE
# ============================================================
image_col = headers["Image"]
rnp_col   = headers["RITNINGSNUMMER_PROJEKT"]

for row_idx in range(2, len(df) + 2):

    if ws.cell(row=row_idx, column=status_col).value == "ERROR":
        continue

    image_name = str(ws.cell(row=row_idx, column=image_col).value).strip()
    rnp_value  = str(ws.cell(row=row_idx, column=rnp_col).value).strip()

    if not image_name or not rnp_value:
        continue

    image_base = (
        image_name
        .replace("_stamp.png", "")
        .replace(".png", "")
        .strip()
    )

    if rnp_value not in image_base:
        ws.cell(row=row_idx, column=rnp_col).fill = RED
        
# ============================================================
# BLAD vs IMAGE VALIDATION
# ============================================================

def extract_digits_from_image(image_name):

    if not image_name:
        return ""

    name = str(image_name).upper()

    name = name.replace("_PDF_STAMP.PNG", "")
    name = name.replace("_STAMP.PNG", "")
    name = name.replace(".PNG", "")

    m = re.search(r"-([0-9]{2,4})$", name)

    return m.group(1) if m else ""


blad_col = headers["BLAD"]
image_col = headers["Image"]

for row_idx in range(2, len(df) + 2):

    if ws.cell(row=row_idx, column=status_col).value == "ERROR":
        continue

    blad_raw = ws.cell(row=row_idx, column=blad_col).value
    blad_value = str(blad_raw).strip() if blad_raw else ""

    image_name = ws.cell(row=row_idx, column=image_col).value

    if blad_value in ("", "0", "00", "000", "0000"):
        continue

    if not image_name:
        continue

    image_digits = extract_digits_from_image(image_name)

    if not image_digits:
        continue

    try:
        if int(blad_value) != int(image_digits):
            ws.cell(row=row_idx, column=blad_col).fill = RED
    except:
        ws.cell(row=row_idx, column=blad_col).fill = RED

# ============================================================
# MASTER VALIDATION LOOP
# ============================================================
for col_idx, col_name in enumerate(df.columns, start=1):

    col_name = col_name.strip()

    if col_name in ("ANDR", "FINAL_REV", "REV_STATUS"):
        continue

    for row_idx in range(2, len(df) + 2):

        if ws.cell(row=row_idx, column=status_col).value == "ERROR":
            continue

        cell = ws.cell(row=row_idx, column=col_idx)
        value = str(cell.value).strip() if cell.value else ""

        if col_name in freetext_labels:
            continue

        if col_name in value_rules:
            if value == "":
                if col_name not in EMPTY_ALLOWED:
                    cell.fill = RED
                continue

            if value not in value_rules[col_name]:
                cell.fill = RED
            continue

        if col_name in pattern_rules:
            if value == "":
                if col_name not in EMPTY_ALLOWED:
                    cell.fill = RED
                continue

            value_nospace = re.sub(r"\s+", "", value)

            if not any(
                re.fullmatch(pat, value_nospace)
                for pat in pattern_rules[col_name]
                if pat
            ):
                cell.fill = RED
            continue

# ============================================================
# SAVE EXCEL
# ============================================================
wb.save(OUTPUT_FILE)

# ============================================================
# CSV EXPORT
# ============================================================
csv_file = str(OUTPUT_FILE).replace(".xlsx", ".csv")

df_csv = pd.read_excel(OUTPUT_FILE, dtype=str, keep_default_na=False)
df_csv.to_csv(csv_file, index=False, encoding="utf-8-sig")

print("\nSTEP-3 VALIDATION COMPLETE")
print("Excel Output:", OUTPUT_FILE)
print("CSV Output:", csv_file)